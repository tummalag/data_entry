from openpyxl import load_workbook
import easygui

workbook_name = 'List.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

# New data to write:
def dataEntry():
    fieldNames = ['Date','Job ID','Job Title','Job Description','Company Name','Client Name']
    fieldVals = easygui.multenterbox('Enter the Data','Job Entry List', fieldNames)
    d, j_id, j_title, j_description,c_name, cli_name   = fieldVals
    row = [fieldVals]
    
    if j_id == '': 
        easygui.msgbox(msg="Data is required", title="Data Error", ok_button="OK")
        return

    for r in range(2,page.max_row+1):
        cell = "{}{}".format('B', r)
        if j_id == page[cell].value:
            easygui.msgbox(msg="Data already exist", title="Data Error", ok_button="OK")
            return
    
    for info in row:
        page.append(info)
        
dataEntry()
while True:
    if easygui.boolbox('Do you have data to Enter??', 'Data', ["YES", "NO "]):
        dataEntry() 
    
    else: break

print("Total Number of records:",page.max_row-1)
wb.save(filename=workbook_name)